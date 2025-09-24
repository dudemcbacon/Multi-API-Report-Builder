"""
Interactive Data Grid Component
Excel-like data viewer and editor using PyQt6 QTableWidget
"""
import logging
from typing import Optional, Dict, Any, List
from pathlib import Path
import polars as pl
from PyQt6.QtWidgets import (
    QWidget, QVBoxLayout, QHBoxLayout, QTableWidget, QTableWidgetItem,
    QPushButton, QLineEdit, QLabel, QComboBox, QSpinBox, QCheckBox,
    QHeaderView, QAbstractItemView, QMenu, QFileDialog, QMessageBox,
    QProgressBar, QFrame, QSplitter, QTextEdit, QTabWidget, QApplication
)
from PyQt6.QtCore import Qt, QThread, pyqtSignal, QTimer
from PyQt6.QtGui import QAction, QFont, QColor, QPalette
import qtawesome as qta

logger = logging.getLogger(__name__)

# Import xlsxwriter for Excel export with proper dynamic array support
try:
    import xlsxwriter
    XLSXWRITER_AVAILABLE = True
except ImportError:
    logger.warning("xlsxwriter not available - Excel export will be disabled")
    xlsxwriter = None
    XLSXWRITER_AVAILABLE = False

# Keep openpyxl as fallback for compatibility
try:
    import openpyxl
    import openpyxl.styles
    OPENPYXL_AVAILABLE = True
except ImportError:
    logger.warning("openpyxl not available - fallback Excel export will be disabled")
    openpyxl = None
    OPENPYXL_AVAILABLE = False

class MultiSheetExportWorker(QThread):
    """Worker thread for multi-sheet Excel export operations"""
    
    export_progress = pyqtSignal(int)
    export_complete = pyqtSignal(str)
    export_error = pyqtSignal(str)
    
    def __init__(self, datasets: Dict[str, pl.DataFrame], file_path: str):
        super().__init__()
        self.datasets = datasets  # Dict with sheet_name: dataframe pairs
        self.file_path = file_path
    
    def run(self):
        try:
            logger.info(f"[MULTI-EXPORT] Starting multi-sheet export to {self.file_path}")
            self.export_progress.emit(10)
            
            # Try xlsxwriter first for better dynamic array support
            if XLSXWRITER_AVAILABLE:
                self._run_xlsxwriter_export()
            elif OPENPYXL_AVAILABLE:
                logger.warning("[MULTI-EXPORT] xlsxwriter not available, falling back to openpyxl")
                self._run_openpyxl_export()
            else:
                self.export_error.emit("Neither xlsxwriter nor openpyxl is available for Excel export")
                return
                
        except Exception as e:
            logger.error(f"[MULTI-EXPORT] Export failed: {e}", exc_info=True)
            self.export_error.emit(f"Export failed: {str(e)}")
    
    def _run_xlsxwriter_export(self):
        """Export using xlsxwriter for proper Excel 365 dynamic array support"""
        logger.info("[MULTI-EXPORT] Using xlsxwriter for export")
        
        # Create workbook with xlsxwriter
        workbook = xlsxwriter.Workbook(self.file_path)
        
        total_sheets = len(self.datasets)
        current_sheet = 0
        
        for sheet_name, df in self.datasets.items():
            if df is None or df.is_empty():
                logger.info(f"[MULTI-EXPORT] Skipping empty sheet: {sheet_name}")
                continue
                
            logger.info(f"[MULTI-EXPORT] Creating sheet: {sheet_name} ({len(df)} rows)")
            
            # Create worksheet
            worksheet = workbook.add_worksheet(sheet_name)
            
            # Convert DataFrame to records
            records = df.to_dicts()
            columns = df.columns
            
            # Create header format
            header_format = workbook.add_format({
                'bold': True,
                'bg_color': '#CCCCCC',
                'border': 1
            })
            
            # Write headers
            for col_idx, column in enumerate(columns):
                worksheet.write(0, col_idx, column, header_format)
            
            # Write data (with special handling for tie-out sheets)
            is_tieout_sheet = "Tie Out" in sheet_name
            for row_idx, record in enumerate(records, 1):  # xlsxwriter uses 0-based indexing
                for col_idx, column in enumerate(columns):
                    value = record.get(column, '')
                    
                    # For tie-out sheets, clear Difference column data to make room for spill formulas
                    if is_tieout_sheet and column == "Difference":
                        # Check if this is the totals row by looking for "Total" in first column
                        first_col_value = record.get(columns[0], '')
                        if str(first_col_value).lower() == "total":
                            # Keep the calculated value for totals row
                            pass  # Don't clear this value
                        else:
                            # Clear data row values to make room for spill formula
                            value = ''
                            logger.debug(f"[MULTI-EXPORT] Cleared Difference value for data row in sheet '{sheet_name}', row {row_idx + 1}")
                    
                    # Sanitize value for Excel compatibility
                    try:
                        if isinstance(value, (list, dict)):
                            if value:  # Only log non-empty complex types
                                logger.debug(f"[MULTI-EXPORT] Converting complex data type in sheet '{sheet_name}', col '{column}': {type(value).__name__}")
                            value = str(value) if value else ''
                        elif value is None:
                            value = ''
                        elif not isinstance(value, (str, int, float, bool)):
                            logger.debug(f"[MULTI-EXPORT] Converting non-standard type in sheet '{sheet_name}', col '{column}': {type(value).__name__}")
                            value = str(value)
                    except Exception as e:
                        logger.warning(f"[MULTI-EXPORT] Error sanitizing value in sheet '{sheet_name}', row {row_idx + 1}, col '{column}': {e}")
                        value = str(value) if value is not None else ''
                    
                    worksheet.write(row_idx, col_idx, value)
            
            # Add Excel formulas for tie-out sheets using xlsxwriter
            if "Tie Out" in sheet_name:
                self._add_xlsxwriter_tieout_formulas(worksheet, workbook, columns, len(records))
            
            # Auto-adjust column widths
            for col_idx, column in enumerate(columns):
                # Calculate max width for this column
                max_length = len(str(column))  # Header length
                for record in records:
                    value = record.get(column, '')
                    max_length = max(max_length, len(str(value)))
                
                # Set column width (capped at 50)
                adjusted_width = min(max_length + 2, 50)
                worksheet.set_column(col_idx, col_idx, adjusted_width)
            
            current_sheet += 1
            progress = 10 + (current_sheet / total_sheets * 80)
            self.export_progress.emit(int(progress))
        
        # Close workbook (automatically saves)
        workbook.close()
        self.export_progress.emit(100)
        
        sheet_names = list(self.datasets.keys())
        self.export_complete.emit(f"Successfully exported {len(sheet_names)} sheets: {', '.join(sheet_names)}")
    
    def _run_openpyxl_export(self):
        """Fallback export using openpyxl (original implementation)"""
        logger.info("[MULTI-EXPORT] Using openpyxl for export (fallback)")
        
        # Create workbook with multiple sheets
        workbook = openpyxl.Workbook()
        
        # Remove default sheet
        if 'Sheet' in workbook.sheetnames:
            workbook.remove(workbook['Sheet'])
        
        total_sheets = len(self.datasets)
        current_sheet = 0
        
        for sheet_name, df in self.datasets.items():
            if df is None or df.is_empty():
                logger.info(f"[MULTI-EXPORT] Skipping empty sheet: {sheet_name}")
                continue
                
            logger.info(f"[MULTI-EXPORT] Creating sheet: {sheet_name} ({len(df)} rows)")
            
            # Create worksheet
            worksheet = workbook.create_sheet(title=sheet_name)
            
            # Convert DataFrame to records for openpyxl
            records = df.to_dicts()
            columns = df.columns
            
            # Write headers
            for col_idx, column in enumerate(columns, 1):
                cell = worksheet.cell(row=1, column=col_idx, value=column)
                cell.font = openpyxl.styles.Font(bold=True)
                cell.fill = openpyxl.styles.PatternFill(start_color="CCCCCC", end_color="CCCCCC", fill_type="solid")
            
            # Write data
            for row_idx, record in enumerate(records, 2):
                for col_idx, column in enumerate(columns, 1):
                    value = record.get(column, '')
                    
                    # Sanitize value for Excel compatibility
                    try:
                        if isinstance(value, (list, dict)):
                            value = str(value) if value else ''
                        elif value is None:
                            value = ''
                        elif not isinstance(value, (str, int, float, bool)):
                            value = str(value)
                    except Exception as e:
                        logger.warning(f"[MULTI-EXPORT] Error sanitizing value: {e}")
                        value = str(value) if value is not None else ''
                    
                    worksheet.cell(row=row_idx, column=col_idx, value=value)
            
            # Auto-adjust column widths
            for column in worksheet.columns:
                max_length = 0
                column_letter = column[0].column_letter
                for cell in column:
                    try:
                        if len(str(cell.value)) > max_length:
                            max_length = len(str(cell.value))
                    except:
                        pass
                adjusted_width = min(max_length + 2, 50)
                worksheet.column_dimensions[column_letter].width = adjusted_width
            
            current_sheet += 1
            progress = 10 + (current_sheet / total_sheets * 80)
            self.export_progress.emit(int(progress))
        
        # Save workbook
        logger.info(f"[MULTI-EXPORT] Saving workbook to {self.file_path}")
        workbook.save(self.file_path)
        self.export_progress.emit(100)
        
        sheet_names = list(self.datasets.keys())
        self.export_complete.emit(f"Successfully exported {len(sheet_names)} sheets: {', '.join(sheet_names)}")
    
    def _add_xlsxwriter_tieout_formulas(self, worksheet, workbook, columns, num_data_rows):
        """Add Excel 365 dynamic array formulas using xlsxwriter for proper spill behavior"""
        try:
            # Find column indices
            diff_col_idx = None
            amount_col_indices = []
            amount_col_names = []
            
            for idx, col in enumerate(columns):
                if col == "Difference":
                    diff_col_idx = idx  # xlsxwriter uses 0-based indexing
                elif "Amount" in col and col != "Difference":
                    amount_col_indices.append(idx)
                    amount_col_names.append(col)
            
            if diff_col_idx is not None and len(amount_col_indices) >= 2:
                # Convert column indices to Excel column letters
                def excel_column_letter(col_num):
                    """Convert column number to Excel letter (0->A, 1->B, etc.)"""
                    result = ""
                    while col_num >= 0:
                        result = chr(65 + col_num % 26) + result
                        col_num = col_num // 26 - 1
                        if col_num < 0:
                            break
                    return result
                
                col1_letter = excel_column_letter(amount_col_indices[0])  # First amount column
                col2_letter = excel_column_letter(amount_col_indices[1])  # Second amount column
                diff_letter = excel_column_letter(diff_col_idx)  # Difference column
                
                # Find totals row by scanning the data
                totals_row = None
                for row_idx in range(1, num_data_rows + 1):  # xlsxwriter 0-based, skip header
                    # We need to check our original data since it's already written
                    # For now, assume last row is totals (this could be improved)
                    pass
                
                # Assume totals is the last row for now
                totals_row = num_data_rows  # Last data row in 0-based indexing
                last_data_row = num_data_rows - 1  # Data ends before totals
                data_range_start = 1  # Start after headers (0-based)
                
                logger.info(f"[XLSXWRITER] Adding formulas to tie-out sheet - Difference column: {diff_letter}, Amount columns: {col1_letter}, {col2_letter}")
                logger.info(f"[XLSXWRITER] Data range: {data_range_start + 1} to {last_data_row + 1}, Totals row: {totals_row + 1}")
                
                # Add dynamic array spill formula for difference calculation (Excel 365 style)
                if last_data_row >= data_range_start:  # Make sure we have data rows
                    # Use Excel's 1-based row numbering in formula
                    diff_formula = f"=ABS({col1_letter}{data_range_start + 1}:{col1_letter}{last_data_row + 1})-ABS({col2_letter}{data_range_start + 1}:{col2_letter}{last_data_row + 1})"
                    
                    # Use xlsxwriter's write_dynamic_array_formula for proper Excel 365 support
                    worksheet.write_dynamic_array_formula(data_range_start, diff_col_idx, data_range_start, diff_col_idx, diff_formula)
                    logger.info(f"[XLSXWRITER] Added dynamic array spill formula in {diff_letter}{data_range_start + 1}: {diff_formula}")
                
                # Add totals formulas using regular write_formula
                if totals_row > data_range_start:
                    # Use Excel's 1-based row numbering in formulas
                    col1_sum_formula = f"=SUMPRODUCT(ABS({col1_letter}{data_range_start + 1}:{col1_letter}{last_data_row + 1}))"
                    col2_sum_formula = f"=SUMPRODUCT(ABS({col2_letter}{data_range_start + 1}:{col2_letter}{last_data_row + 1}))"
                    
                    # Write totals formulas
                    worksheet.write_formula(totals_row, amount_col_indices[0], col1_sum_formula)
                    worksheet.write_formula(totals_row, amount_col_indices[1], col2_sum_formula)
                    
                    # Difference in totals row - reference the actual SUM cells
                    diff_total_formula = f"=ABS({col1_letter}{totals_row + 1})-ABS({col2_letter}{totals_row + 1})"
                    worksheet.write_formula(totals_row, diff_col_idx, diff_total_formula)
                    
                    logger.info(f"[XLSXWRITER] Added totals formulas - {amount_col_names[0]}: {col1_sum_formula}, {amount_col_names[1]}: {col2_sum_formula}")
                    logger.info(f"[XLSXWRITER] Added totals difference formula: {diff_total_formula}")
                    
        except Exception as e:
            logger.error(f"[XLSXWRITER] Error adding tie-out formulas: {e}", exc_info=True)
    
    def _add_tieout_formulas(self, worksheet, columns, num_data_rows):
        """Add Excel spill formulas to tie-out sheets for dynamic calculations"""
        try:
            # Find column indices
            diff_col_idx = None
            amount_col_indices = []
            amount_col_names = []
            
            for idx, col in enumerate(columns):
                if col == "Difference":
                    diff_col_idx = idx + 1  # Excel is 1-based
                elif "Amount" in col and col != "Difference":
                    amount_col_indices.append(idx + 1)
                    amount_col_names.append(col)
            
            if diff_col_idx and len(amount_col_indices) >= 2:
                # Convert column indices to Excel column letters
                def excel_column_letter(col_num):
                    """Convert column number to Excel letter (1->A, 2->B, etc.)"""
                    result = ""
                    while col_num > 0:
                        col_num -= 1
                        result = chr(65 + col_num % 26) + result
                        col_num //= 26
                    return result
                
                col1_letter = excel_column_letter(amount_col_indices[0])  # First amount column
                col2_letter = excel_column_letter(amount_col_indices[1])  # Second amount column
                diff_letter = excel_column_letter(diff_col_idx)  # Difference column
                
                # Determine data range (excluding totals row)
                # Find totals row by looking for "Total" in first column
                totals_row = None
                for row_idx in range(2, num_data_rows + 2):  # Start from row 2 (after headers)
                    cell_value = worksheet.cell(row=row_idx, column=1).value
                    if cell_value and str(cell_value).lower() == "total":
                        totals_row = row_idx
                        break
                
                if totals_row:
                    last_data_row = totals_row - 1  # Data ends before totals
                else:
                    last_data_row = num_data_rows  # No totals row found
                
                data_range_start = 2  # Start after headers
                
                logger.info(f"[MULTI-EXPORT] Adding formulas to tie-out sheet - Difference column: {diff_letter}, Amount columns: {col1_letter}, {col2_letter}")
                logger.info(f"[MULTI-EXPORT] Data range: {data_range_start} to {last_data_row}, Totals row: {totals_row}")
                
                # Ensure difference column is clear for data rows (should already be cleared above)
                # This is a safety check to make sure we have clean cells for the spill formula
                for row_idx in range(data_range_start, last_data_row + 1):
                    current_value = worksheet.cell(row=row_idx, column=diff_col_idx).value
                    if current_value is not None and current_value != '':
                        worksheet.cell(row=row_idx, column=diff_col_idx, value='')
                        logger.debug(f"[MULTI-EXPORT] Safety cleared cell {row_idx}, {diff_col_idx}")
                
                # Add spill formula ONLY in the first data cell using ArrayFormula to prevent @ symbols
                if last_data_row >= data_range_start:  # Make sure we have data rows
                    # Create formula without @ symbols using ArrayFormula class
                    diff_formula = f"ABS({col1_letter}{data_range_start}:{col1_letter}{last_data_row})-ABS({col2_letter}{data_range_start}:{col2_letter}{last_data_row})"
                    cell_ref = f"{diff_letter}{data_range_start}"
                    
                    # Try entering formula with explicit data type to avoid @ symbol processing
                    cell = worksheet.cell(row=data_range_start, column=diff_col_idx)
                    formula_string = f"=ABS({col1_letter}{data_range_start}:{col1_letter}{last_data_row})-ABS({col2_letter}{data_range_start}:{col2_letter}{last_data_row})"
                    # Set value first, then data type
                    cell.value = formula_string
                    cell.data_type = 'f'  # 'f' for formula
                    logger.info(f"[MULTI-EXPORT] Added formula with explicit type in {cell_ref}: {formula_string}")
                
                # Update totals row with SUM formulas (if exists)
                if totals_row and totals_row > data_range_start:
                    # Replace amount totals with SUM formulas for Excel 365 (no curly brackets needed)
                    col1_sum_formula = f"SUMPRODUCT(ABS({col1_letter}{data_range_start}:{col1_letter}{last_data_row}))"
                    col2_sum_formula = f"SUMPRODUCT(ABS({col2_letter}{data_range_start}:{col2_letter}{last_data_row}))"
                    
                    # Set formula cells with explicit data type to avoid @ symbols
                    col1_cell = worksheet.cell(row=totals_row, column=amount_col_indices[0])
                    col2_cell = worksheet.cell(row=totals_row, column=amount_col_indices[1])
                    
                    col1_formula_string = f"={col1_sum_formula}"
                    col2_formula_string = f"={col2_sum_formula}"
                    
                    col1_cell.value = col1_formula_string
                    col1_cell.data_type = 'f'  # 'f' for formula
                    col2_cell.value = col2_formula_string
                    col2_cell.data_type = 'f'  # 'f' for formula
                    
                    logger.info(f"[MULTI-EXPORT] Added totals formulas with explicit type - {amount_col_names[0]}: {col1_formula_string}, {amount_col_names[1]}: {col2_formula_string}")
                    
                    # Difference in totals row - reference the actual SUM cells (simple formula, no array needed)
                    diff_total_formula = f"ABS({col1_letter}{totals_row})-ABS({col2_letter}{totals_row})"
                    diff_cell = worksheet.cell(row=totals_row, column=diff_col_idx)
                    diff_total_formula_string = f"={diff_total_formula}"
                    diff_cell.value = diff_total_formula_string
                    diff_cell.data_type = 'f'  # 'f' for formula
                    
                    logger.info(f"[MULTI-EXPORT] Added totals difference formula with explicit type: {diff_total_formula_string}")
                    
        except Exception as e:
            logger.error(f"[MULTI-EXPORT] Error adding tie-out formulas: {e}", exc_info=True)


class DataExportWorker(QThread):
    """Worker thread for single-sheet data export operations"""
    
    export_progress = pyqtSignal(int)
    export_complete = pyqtSignal(str)
    export_error = pyqtSignal(str)
    
    def __init__(self, dataframe: pl.DataFrame, file_path: str, format_type: str = "excel"):
        super().__init__()
        self.dataframe = dataframe
        self.file_path = file_path
        self.format_type = format_type
    
    def run(self):
        try:
            logger.info(f"[EXPORT] Starting export to {self.file_path}")
            self.export_progress.emit(10)
            
            if self.format_type == "excel":
                # Try Polars native Excel export first (which uses xlsxwriter under the hood)
                try:
                    self.export_progress.emit(30)
                    
                    # Use Polars native write_excel method
                    logger.info("[EXPORT] Using Polars native Excel export")
                    self.dataframe.write_excel(
                        self.file_path,
                        worksheet="Data",
                        autofit=True,
                        table_style="Table Style Light 9",
                        float_precision=2
                    )
                    
                    self.export_progress.emit(90)
                    
                except ImportError as e:
                    logger.warning(f"[EXPORT] Polars Excel export not available: {e}")
                    
                    # Try xlsxwriter fallback
                    if XLSXWRITER_AVAILABLE:
                        logger.info("[EXPORT] Falling back to xlsxwriter manual export")
                        self._export_with_xlsxwriter()
                    elif OPENPYXL_AVAILABLE:
                        logger.info("[EXPORT] Falling back to openpyxl manual export")
                        self._export_with_openpyxl()
                    else:
                        logger.warning("[EXPORT] No Excel libraries available, falling back to CSV export")
                        # Fallback to CSV export
                        csv_path = self.file_path.replace('.xlsx', '.csv')
                        self.export_progress.emit(50)
                        self.dataframe.write_csv(csv_path)
                        self.export_progress.emit(100)
                        self.export_complete.emit(csv_path)
                        return
                
            elif self.format_type == "csv":
                self.dataframe.write_csv(self.file_path)
                self.export_progress.emit(90)
            
            self.export_progress.emit(100)
            self.export_complete.emit(self.file_path)
            logger.info(f"[EXPORT] Export completed: {self.file_path}")
            
        except Exception as e:
            logger.error(f"[EXPORT] Export failed: {e}")
            self.export_error.emit(str(e))
    
    def _export_with_xlsxwriter(self):
        """Export using xlsxwriter"""
        self.export_progress.emit(40)
        
        # Create workbook and worksheet
        workbook = xlsxwriter.Workbook(self.file_path)
        worksheet = workbook.add_worksheet('Data')
        
        # Create header format
        header_format = workbook.add_format({
            'bold': True,
            'bg_color': '#CCCCCC',
            'border': 1
        })
        
        self.export_progress.emit(50)
        
        # Write headers
        headers = self.dataframe.columns
        for col_idx, header in enumerate(headers):
            worksheet.write(0, col_idx, header, header_format)
        
        self.export_progress.emit(60)
        
        # Write data rows efficiently
        for row_idx, row in enumerate(self.dataframe.iter_rows(), 1):
            for col_idx, value in enumerate(row):
                # Handle None/null values
                if value is None:
                    value = ""
                worksheet.write(row_idx, col_idx, value)
            
            # Update progress for large datasets
            if row_idx % 1000 == 0:
                progress = 60 + ((row_idx - 1) / len(self.dataframe)) * 20
                self.export_progress.emit(int(progress))
        
        self.export_progress.emit(80)
        
        # Auto-adjust column widths
        for col_idx, header in enumerate(headers):
            # Calculate max width for this column
            max_length = len(str(header))  # Header length
            for row in self.dataframe.iter_rows():
                value = row[col_idx]
                max_length = max(max_length, len(str(value)))
            
            # Set column width (capped at 50)
            adjusted_width = min(max_length + 2, 50)
            worksheet.set_column(col_idx, col_idx, adjusted_width)
        
        self.export_progress.emit(85)
        
        # Close workbook (automatically saves)
        workbook.close()
        self.export_progress.emit(90)
    
    def _export_with_openpyxl(self):
        """Export using openpyxl as fallback"""
        self.export_progress.emit(40)
        
        # Create workbook and worksheet
        workbook = openpyxl.Workbook()
        worksheet = workbook.active
        worksheet.title = 'Data'
        
        self.export_progress.emit(50)
        
        # Write headers
        headers = self.dataframe.columns
        for col_idx, header in enumerate(headers, 1):
            cell = worksheet.cell(row=1, column=col_idx, value=header)
            cell.font = openpyxl.styles.Font(bold=True)
            cell.fill = openpyxl.styles.PatternFill(start_color="CCCCCC", end_color="CCCCCC", fill_type="solid")
        
        self.export_progress.emit(60)
        
        # Write data rows efficiently
        for row_idx, row in enumerate(self.dataframe.iter_rows(), 2):
            for col_idx, value in enumerate(row, 1):
                # Handle None/null values
                if value is None:
                    value = ""
                worksheet.cell(row=row_idx, column=col_idx, value=value)
            
            # Update progress for large datasets
            if row_idx % 1000 == 0:
                progress = 60 + ((row_idx - 2) / len(self.dataframe)) * 20
                self.export_progress.emit(int(progress))
        
        self.export_progress.emit(80)
        
        # Auto-adjust column widths
        for column in worksheet.columns:
            max_length = 0
            column_letter = column[0].column_letter
            for cell in column:
                try:
                    if len(str(cell.value)) > max_length:
                        max_length = len(str(cell.value))
                except:
                    pass
            adjusted_width = min(max_length + 2, 50)
            worksheet.column_dimensions[column_letter].width = adjusted_width
        
        self.export_progress.emit(85)
        
        # Save workbook
        workbook.save(self.file_path)
        workbook.close()
        self.export_progress.emit(90)

class DataFilterWidget(QWidget):
    """Widget for filtering and searching data"""
    
    filter_changed = pyqtSignal(str, str, str)  # column, operator, value
    
    def __init__(self):
        super().__init__()
        self.setup_ui()
    
    def setup_ui(self):
        layout = QHBoxLayout(self)
        layout.setContentsMargins(5, 5, 5, 5)
        
        # Search box
        layout.addWidget(QLabel("Search:"))
        self.search_box = QLineEdit()
        self.search_box.setPlaceholderText("Search all columns...")
        self.search_box.textChanged.connect(self.on_search_changed)
        layout.addWidget(self.search_box)
        
        layout.addWidget(QLabel("|"))
        
        # Column filter
        layout.addWidget(QLabel("Filter:"))
        self.column_combo = QComboBox()
        self.column_combo.setMinimumWidth(120)
        layout.addWidget(self.column_combo)
        
        # Operator
        self.operator_combo = QComboBox()
        self.operator_combo.addItems(["Contains", "Equals", "Starts with", "Ends with", "Greater than", "Less than"])
        layout.addWidget(self.operator_combo)
        
        # Value
        self.value_edit = QLineEdit()
        self.value_edit.setPlaceholderText("Filter value...")
        layout.addWidget(self.value_edit)
        
        # Apply filter button
        filter_btn = QPushButton("Apply")
        filter_btn.setIcon(qta.icon('fa5s.filter'))
        filter_btn.clicked.connect(self.apply_filter)
        layout.addWidget(filter_btn)
        
        # Clear filter button
        clear_btn = QPushButton("Clear")
        clear_btn.setIcon(qta.icon('fa5s.times'))
        clear_btn.clicked.connect(self.clear_filter)
        layout.addWidget(clear_btn)
        
        layout.addStretch()
    
    def set_columns(self, columns: List[str]):
        """Set available columns for filtering"""
        self.column_combo.clear()
        self.column_combo.addItems(["All Columns"] + columns)
    
    def on_search_changed(self):
        """Handle search box text change"""
        self.filter_changed.emit("", "search", self.search_box.text())
    
    def apply_filter(self):
        """Apply column-specific filter"""
        column = self.column_combo.currentText()
        operator = self.operator_combo.currentText()
        value = self.value_edit.text()
        self.filter_changed.emit(column, operator, value)
    
    def clear_filter(self):
        """Clear all filters"""
        self.search_box.clear()
        self.value_edit.clear()
        self.column_combo.setCurrentIndex(0)
        self.filter_changed.emit("", "clear", "")

class DataStatsWidget(QWidget):
    """Widget showing data statistics"""
    
    def __init__(self):
        super().__init__()
        self.setup_ui()
    
    def setup_ui(self):
        layout = QHBoxLayout(self)
        layout.setContentsMargins(5, 5, 5, 5)
        
        self.stats_label = QLabel("No data loaded")
        self.stats_label.setStyleSheet("color: #666; font-size: 12px;")
        layout.addWidget(self.stats_label)
        
        layout.addStretch()
        
        # Refresh button
        refresh_btn = QPushButton("Refresh")
        refresh_btn.setIcon(qta.icon('fa5s.sync'))
        refresh_btn.setToolTip("Refresh data from source")
        layout.addWidget(refresh_btn)
    
    def update_stats(self, total_rows: int, filtered_rows: int, columns: int):
        """Update statistics display"""
        if filtered_rows < total_rows:
            self.stats_label.setText(f"Showing {filtered_rows:,} of {total_rows:,} rows, {columns} columns")
        else:
            self.stats_label.setText(f"{total_rows:,} rows, {columns} columns")

class InteractiveDataGrid(QWidget):
    """
    Excel-like interactive data grid component
    Features: sorting, filtering, editing, export, search
    """
    
    data_changed = pyqtSignal()
    
    def __init__(self, dataframe: Optional[pl.DataFrame] = None, title: str = "Data"):
        super().__init__()
        self.original_dataframe = dataframe
        self.filtered_dataframe = dataframe
        self.title = title
        self.current_filters = {}
        self.export_worker = None
        self.multi_export_worker = None
        
        self.setup_ui()
        
        if dataframe is not None:
            self.load_data(dataframe)
    
    def __del__(self):
        """Ensure worker threads are properly cleaned up"""
        self.cleanup_workers()
    
    def cleanup_workers(self):
        """Clean up any running worker threads"""
        if self.export_worker and self.export_worker.isRunning():
            self.export_worker.quit()
            self.export_worker.wait(1000)  # Wait up to 1 second
            self.export_worker.deleteLater()
            self.export_worker = None
        
        if self.multi_export_worker and self.multi_export_worker.isRunning():
            self.multi_export_worker.quit()
            self.multi_export_worker.wait(1000)  # Wait up to 1 second
            self.multi_export_worker.deleteLater()
            self.multi_export_worker = None
    
    def setup_ui(self):
        """Setup the data grid UI"""
        layout = QVBoxLayout(self)
        layout.setContentsMargins(0, 0, 0, 0)
        
        # Top toolbar
        toolbar_layout = QHBoxLayout()
        
        # Title
        title_label = QLabel(self.title)
        title_label.setStyleSheet("font-size: 14px; font-weight: bold; color: #333;")
        toolbar_layout.addWidget(title_label)
        
        toolbar_layout.addStretch()
        
        # Visualize button
        visualize_btn = QPushButton("Visualize")
        visualize_btn.setIcon(qta.icon('fa5s.chart-bar'))
        visualize_btn.clicked.connect(self.visualize_data)
        visualize_btn.setToolTip("Send this data to the Visualization tab")
        toolbar_layout.addWidget(visualize_btn)
        
        # Export button
        export_btn = QPushButton("Export")
        export_btn.setIcon(qta.icon('fa5s.download'))
        export_btn.clicked.connect(self.export_data)
        toolbar_layout.addWidget(export_btn)
        
        # Add to columns button (for adding calculated columns)
        add_col_btn = QPushButton("Add Column")
        add_col_btn.setIcon(qta.icon('fa5s.plus'))
        add_col_btn.clicked.connect(self.add_calculated_column)
        toolbar_layout.addWidget(add_col_btn)
        
        layout.addLayout(toolbar_layout)
        
        # Filter widget
        self.filter_widget = DataFilterWidget()
        self.filter_widget.filter_changed.connect(self.apply_filter)
        layout.addWidget(self.filter_widget)
        
        # Main content with splitter
        splitter = QSplitter(Qt.Orientation.Vertical)
        layout.addWidget(splitter)
        
        # Data table
        self.table = QTableWidget()
        self.table.setSortingEnabled(True)
        self.table.setAlternatingRowColors(True)
        self.table.setSelectionBehavior(QAbstractItemView.SelectionBehavior.SelectRows)
        self.table.horizontalHeader().setStretchLastSection(True)
        self.table.setContextMenuPolicy(Qt.ContextMenuPolicy.CustomContextMenu)
        self.table.customContextMenuRequested.connect(self.show_context_menu)
        
        # Enable editing
        self.table.setEditTriggers(QAbstractItemView.EditTrigger.DoubleClicked | 
                                  QAbstractItemView.EditTrigger.EditKeyPressed)
        
        splitter.addWidget(self.table)
        
        # Bottom info panel (collapsible)
        info_widget = QWidget()
        info_layout = QVBoxLayout(info_widget)
        
        # Stats widget
        self.stats_widget = DataStatsWidget()
        info_layout.addWidget(self.stats_widget)
        
        # Progress bar (hidden by default)
        self.progress_bar = QProgressBar()
        self.progress_bar.setVisible(False)
        info_layout.addWidget(self.progress_bar)
        
        splitter.addWidget(info_widget)
        splitter.setSizes([600, 100])
    
    def load_data(self, dataframe: pl.DataFrame, title: str = None):
        """Load data into the grid"""
        logger.info(f"[DATA-GRID] Loading {len(dataframe)} rows, {len(dataframe.columns)} columns")
        
        self.original_dataframe = dataframe
        self.filtered_dataframe = dataframe
        
        if title:
            self.title = title
        
        # Update filter widget columns
        self.filter_widget.set_columns(dataframe.columns)
        
        # Update table
        self._populate_table(dataframe)
        
        # Update stats
        self.stats_widget.update_stats(len(dataframe), len(dataframe), len(dataframe.columns))
        
        logger.info("[DATA-GRID] Data loaded successfully")
    
    def _populate_table(self, dataframe: pl.DataFrame):
        """Populate the table widget with data"""
        if dataframe is None or len(dataframe) == 0:
            self.table.setRowCount(0)
            self.table.setColumnCount(0)
            return
        
        # Set table dimensions
        self.table.setRowCount(len(dataframe))
        self.table.setColumnCount(len(dataframe.columns))
        self.table.setHorizontalHeaderLabels(dataframe.columns)
        
        # Populate data
        for row_idx in range(len(dataframe)):
            for col_idx, column in enumerate(dataframe.columns):
                value = dataframe[column][row_idx]
                
                # Handle None/null values
                display_value = "" if value is None else str(value)
                
                item = QTableWidgetItem(display_value)
                
                # Make numeric columns right-aligned
                if dataframe[column].dtype in [pl.Int64, pl.Float64, pl.Int32, pl.Float32]:
                    item.setTextAlignment(Qt.AlignmentFlag.AlignRight | Qt.AlignmentFlag.AlignVCenter)
                
                self.table.setItem(row_idx, col_idx, item)
        
        # Auto-resize columns to content
        self.table.resizeColumnsToContents()
        
        # Limit column width
        header = self.table.horizontalHeader()
        for i in range(len(dataframe.columns)):
            width = header.sectionSize(i)
            if width > 200:
                header.resizeSection(i, 200)
    
    def apply_filter(self, column: str, operator: str, value: str):
        """Apply filter to the data"""
        if self.original_dataframe is None or self.original_dataframe.is_empty():
            return
        
        logger.info(f"[DATA-GRID] Applying filter: {column} {operator} '{value}'")
        
        try:
            df = self.original_dataframe
            
            if operator == "clear" or not value:
                # Clear all filters
                self.filtered_dataframe = self.original_dataframe
                self.current_filters.clear()
            elif operator == "search":
                # Global search across all columns
                if value:
                    # Create a boolean mask for any column containing the search term
                    mask = pl.lit(False)
                    for col in df.columns:
                        # Convert to string and search (case-insensitive)
                        mask = mask | df[col].cast(pl.Utf8).str.to_lowercase().str.contains(value.lower(), literal=True)
                    
                    self.filtered_dataframe = df.filter(mask)
                else:
                    self.filtered_dataframe = df
            else:
                # Column-specific filter
                if column and column != "All Columns":
                    if operator == "Contains":
                        mask = df[column].cast(pl.Utf8).str.to_lowercase().str.contains(value.lower(), literal=True)
                    elif operator == "Equals":
                        mask = df[column].cast(pl.Utf8).str.to_lowercase() == value.lower()
                    elif operator == "Starts with":
                        mask = df[column].cast(pl.Utf8).str.to_lowercase().str.starts_with(value.lower())
                    elif operator == "Ends with":
                        mask = df[column].cast(pl.Utf8).str.to_lowercase().str.ends_with(value.lower())
                    elif operator == "Greater than":
                        try:
                            mask = df[column] > float(value)
                        except ValueError:
                            mask = df[column].cast(pl.Utf8) > value
                    elif operator == "Less than":
                        try:
                            mask = df[column] < float(value)
                        except ValueError:
                            mask = df[column].cast(pl.Utf8) < value
                    
                    self.filtered_dataframe = df.filter(mask)
                    self.current_filters[column] = (operator, value)
            
            # Update table display
            self._populate_table(self.filtered_dataframe)
            
            # Update stats
            self.stats_widget.update_stats(
                len(self.original_dataframe), 
                len(self.filtered_dataframe), 
                len(self.original_dataframe.columns)
            )
            
            logger.info(f"[DATA-GRID] Filter applied: {len(self.filtered_dataframe)} rows remaining")
            
        except Exception as e:
            logger.error(f"[DATA-GRID] Filter error: {e}")
            QMessageBox.warning(self, "Filter Error", f"Error applying filter: {e}")
    
    def show_context_menu(self, position):
        """Show context menu for table operations"""
        if self.table.itemAt(position) is None:
            return
        
        menu = QMenu(self)
        
        # Copy cell
        copy_action = QAction("Copy Cell", self)
        copy_action.setIcon(qta.icon('fa5s.copy'))
        copy_action.triggered.connect(self.copy_cell)
        menu.addAction(copy_action)
        
        # Copy row
        copy_row_action = QAction("Copy Row", self)
        copy_row_action.triggered.connect(self.copy_row)
        menu.addAction(copy_row_action)
        
        menu.addSeparator()
        
        # Visualize data
        visualize_action = QAction("Send to Visualization", self)
        visualize_action.setIcon(qta.icon('fa5s.chart-bar'))
        visualize_action.triggered.connect(self.visualize_data)
        visualize_action.setToolTip("Send this data to the Visualization tab")
        menu.addAction(visualize_action)
        
        # Export selected
        export_selected_action = QAction("Export Selected Rows", self)
        export_selected_action.setIcon(qta.icon('fa5s.download'))
        export_selected_action.triggered.connect(self.export_selected)
        menu.addAction(export_selected_action)
        
        menu.exec(self.table.mapToGlobal(position))
    
    def copy_cell(self):
        """Copy current cell to clipboard"""
        current_item = self.table.currentItem()
        if current_item:
            QApplication.clipboard().setText(current_item.text())
    
    def copy_row(self):
        """Copy current row to clipboard"""
        current_row = self.table.currentRow()
        if current_row >= 0:
            row_data = []
            for col in range(self.table.columnCount()):
                item = self.table.item(current_row, col)
                row_data.append(item.text() if item else "")
            QApplication.clipboard().setText("\t".join(row_data))
    
    def export_selected(self):
        """Export selected rows"""
        selected_rows = set()
        for item in self.table.selectedItems():
            selected_rows.add(item.row())
        
        if not selected_rows:
            QMessageBox.information(self, "No Selection", "Please select rows to export.")
            return
        
        # Create dataframe with selected rows
        selected_indices = sorted(list(selected_rows))
        if self.filtered_dataframe is not None and not self.filtered_dataframe.is_empty():
            selected_df = self.filtered_dataframe[selected_indices]
            self._export_dataframe(selected_df, f"{self.title}_selected")
    
    def export_data(self):
        """Export current data (with filters applied)"""
        if self.filtered_dataframe is None or self.filtered_dataframe.is_empty():
            QMessageBox.information(self, "No Data", "No data to export.")
            return
        
        self._export_dataframe(self.filtered_dataframe, self.title)
    
    def export_multi_sheet(self, datasets: Dict[str, pl.DataFrame], default_name: str):
        """Export multiple datasets to a single Excel workbook with multiple sheets"""
        if not XLSXWRITER_AVAILABLE and not OPENPYXL_AVAILABLE:
            QMessageBox.warning(
                self, 
                "Export Unavailable", 
                "Multi-sheet Excel export requires xlsxwriter or openpyxl library.\nPlease install with: pip install xlsxwriter"
            )
            return
        
        # Get file path from user
        file_path, file_type = QFileDialog.getSaveFileName(
            self,
            "Export Multi-Sheet Workbook",
            f"{default_name}.xlsx",
            "Excel files (*.xlsx);;All files (*.*)"
        )
        
        if file_path:
            # Ensure .xlsx extension
            if not file_path.endswith('.xlsx'):
                file_path += '.xlsx'
            
            # Filter out None/empty datasets
            valid_datasets = {name: df for name, df in datasets.items() 
                            if df is not None and not df.is_empty()}
            
            if not valid_datasets:
                QMessageBox.information(self, "No Data", "No data available to export.")
                return
            
            # Clean up any existing worker before starting new one
            if self.multi_export_worker and self.multi_export_worker.isRunning():
                self.multi_export_worker.quit()
                self.multi_export_worker.wait(1000)
                self.multi_export_worker.deleteLater()
            
            # Start export in worker thread
            self.multi_export_worker = MultiSheetExportWorker(valid_datasets, file_path)
            self.multi_export_worker.export_progress.connect(self.progress_bar.setValue, Qt.ConnectionType.QueuedConnection)
            self.multi_export_worker.export_complete.connect(self.on_export_complete, Qt.ConnectionType.QueuedConnection)
            self.multi_export_worker.export_error.connect(self.on_export_error, Qt.ConnectionType.QueuedConnection)
            self.multi_export_worker.finished.connect(self.on_multi_export_finished, Qt.ConnectionType.QueuedConnection)
            
            # Show progress bar and start export
            self.progress_bar.setVisible(True)
            self.progress_bar.setValue(0)
            self.multi_export_worker.start()
    
    def _export_dataframe(self, dataframe: pl.DataFrame, default_name: str):
        """Internal method to export dataframe"""
        # Check if Excel export is available
        if not XLSXWRITER_AVAILABLE and not OPENPYXL_AVAILABLE:
            # Only offer CSV if no Excel libraries are available
            file_path, file_type = QFileDialog.getSaveFileName(
                self,
                "Export Data (Excel dependencies not available - CSV only)",
                f"{default_name}.csv",
                "CSV files (*.csv);;All files (*.*)"
            )
        else:
            file_path, file_type = QFileDialog.getSaveFileName(
                self,
                "Export Data",
                f"{default_name}.xlsx",
                "Excel files (*.xlsx);;CSV files (*.csv);;All files (*.*)"
            )
        
        if file_path:
            # Determine format from extension
            format_type = "excel" if file_path.endswith('.xlsx') else "csv"
            
            # Clean up any existing worker before starting new one
            if self.export_worker and self.export_worker.isRunning():
                self.export_worker.quit()
                self.export_worker.wait(1000)
                self.export_worker.deleteLater()
            
            # Start export in worker thread
            self.export_worker = DataExportWorker(dataframe, file_path, format_type)
            self.export_worker.export_progress.connect(self.progress_bar.setValue, Qt.ConnectionType.QueuedConnection)
            self.export_worker.export_complete.connect(self.on_export_complete, Qt.ConnectionType.QueuedConnection)
            self.export_worker.export_error.connect(self.on_export_error, Qt.ConnectionType.QueuedConnection)
            self.export_worker.finished.connect(self.on_export_finished, Qt.ConnectionType.QueuedConnection)
            
            self.progress_bar.setVisible(True)
            self.progress_bar.setValue(0)
            self.export_worker.start()
    
    def on_export_complete(self, file_path: str):
        """Handle export completion"""
        self.progress_bar.setVisible(False)
        
        # Check if this was a fallback CSV export
        message = f"Data exported successfully to:\n{file_path}"
        if file_path.endswith('.csv') and not OPENPYXL_AVAILABLE:
            message += "\n\nNote: Excel export not available (missing openpyxl), exported as CSV instead."
        
        QMessageBox.information(self, "Export Complete", message)
    
    def on_export_error(self, error: str):
        """Handle export error"""
        self.progress_bar.setVisible(False)
        QMessageBox.critical(self, "Export Error", f"Export failed:\n{error}")
    
    def on_multi_export_finished(self):
        """Handle multi-sheet export thread completion and cleanup"""
        if hasattr(self, 'multi_export_worker') and self.multi_export_worker:
            self.multi_export_worker.deleteLater()
            self.multi_export_worker = None
    
    def on_export_finished(self):
        """Handle single-sheet export thread completion and cleanup"""
        if hasattr(self, 'export_worker') and self.export_worker:
            self.export_worker.deleteLater()
            self.export_worker = None
    
    def add_calculated_column(self):
        """Add a new calculated column"""
        # This would open a dialog for creating calculated columns
        # For now, just show a placeholder message
        QMessageBox.information(self, "Feature Coming Soon", 
                              "Calculated columns feature will be implemented in a future update.")
    
    def visualize_data(self):
        """Send current data to the visualization tab"""
        try:
            # Get current data
            current_data = self.get_current_data()
            if current_data is None or len(current_data) == 0:
                QMessageBox.information(self, "No Data", "No data available to visualize.")
                return
            
            # Find the main window
            main_window = self
            while main_window.parent() is not None:
                main_window = main_window.parent()
                if hasattr(main_window, 'visualization_tab'):
                    break
            
            if not hasattr(main_window, 'visualization_tab'):
                QMessageBox.warning(self, "Visualization Not Available", 
                                  "Visualization tab is not available.")
                return
            
            # Get the dataset name (use the grid's title)
            dataset_name = self.title
            
            # Add the dataset to visualization tab (it will be updated if it already exists)
            main_window.visualization_tab.add_dataset_from_grid(
                dataset_name, 
                current_data, 
                self
            )
            
            # Switch to visualization tab
            if hasattr(main_window, 'main_tabs'):
                # Find the visualization tab index
                for i in range(main_window.main_tabs.count()):
                    if main_window.main_tabs.tabText(i) == "Visualization":
                        main_window.main_tabs.setCurrentIndex(i)
                        break
                
                # Auto-select the dataset
                main_window.visualization_tab.visualize_dataset(dataset_name)
            
            logger.info(f"[DATA-GRID] Sent data to visualization: {dataset_name} ({len(current_data)} rows)")
            
        except Exception as e:
            logger.error(f"[DATA-GRID] Error sending data to visualization: {e}")
            QMessageBox.critical(self, "Visualization Error", f"Failed to send data to visualization: {str(e)}")
    
    def get_current_data(self) -> Optional[pl.DataFrame]:
        """Get the currently displayed (filtered) data"""
        return self.filtered_dataframe
    
    def refresh_data(self):
        """Refresh data from source - to be implemented by parent"""
        self.data_changed.emit()

